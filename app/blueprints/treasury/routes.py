"""Treasury blueprint routes — transactions, tags, CERFA, and association config."""

import io
import logging
from datetime import date
from decimal import Decimal

from flask import Response, abort, flash, redirect, render_template, request, url_for

logger = logging.getLogger(__name__)
from flask_login import current_user
from sqlalchemy.orm import selectinload

from app.blueprints.treasury import bp
from app.blueprints.treasury.forms import AssociationConfigForm, TagForm, TransactionForm
from app.decorators import bureau_required
from app.extensions import db
from app.models.config import AssociationConfig
from app.models.treasury import Tag, Transaction, TransactionSource, TransactionType

# ---------------------------------------------------------------------------
# Transaction list / bilan
# ---------------------------------------------------------------------------


@bp.route("/")
@bureau_required
def list_transactions():
    """Render transactions with filters, running balance, and pagination."""
    type_filter = request.args.get("type", "")
    source_filter = request.args.get("source", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    page = request.args.get("page", 1, type=int)

    # Build filter conditions once, reused for totals and pagination
    filters = []
    if type_filter in {e.value for e in TransactionType}:
        filters.append(Transaction.type == type_filter)
    if source_filter in {e.value for e in TransactionSource}:
        filters.append(Transaction.source == source_filter)
    if date_from:
        try:
            filters.append(Transaction.date >= date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            filters.append(Transaction.date <= date.fromisoformat(date_to))
        except ValueError:
            pass

    # Compute totals over the full filtered set via SQL SUM (not in Python)
    total_income = db.session.scalar(
        db.select(db.func.sum(Transaction.amount)).where(
            Transaction.type == TransactionType.INCOME.value, *filters
        )
    ) or Decimal("0")
    total_expense = db.session.scalar(
        db.select(db.func.sum(Transaction.amount)).where(
            Transaction.type == TransactionType.EXPENSE.value, *filters
        )
    ) or Decimal("0")
    balance = total_income - total_expense

    stmt = (
        db.select(Transaction)
        .options(selectinload(Transaction.tags))
        .where(*filters)
        .order_by(Transaction.date.desc(), Transaction.created_at.desc())
    )
    pagination = db.paginate(stmt, page=page, per_page=25, error_out=False)

    return render_template(
        "treasury/list.html",
        transactions=pagination.items,
        pagination=pagination,
        total_income=total_income,
        total_expense=total_expense,
        balance=balance,
        type_filter=type_filter,
        source_filter=source_filter,
        date_from=date_from,
        date_to=date_to,
        TransactionType=TransactionType,
        TransactionSource=TransactionSource,
    )


# ---------------------------------------------------------------------------
# Transaction detail
# ---------------------------------------------------------------------------


@bp.route("/<int:transaction_id>")
@bureau_required
def detail(transaction_id: int):
    """Render the detail page for a transaction."""
    transaction = db.session.get(
        Transaction,
        transaction_id,
        options=[
            selectinload(Transaction.created_by),
            selectinload(Transaction.tags),
        ],
    )
    if transaction is None:
        abort(404)
    return render_template(
        "treasury/detail.html",
        transaction=transaction,
        TransactionType=TransactionType,
        TransactionSource=TransactionSource,
    )


# ---------------------------------------------------------------------------
# Create transaction — bureau only
# ---------------------------------------------------------------------------


@bp.route("/new", methods=["GET", "POST"])
@bureau_required
def create():
    """Create a new financial transaction."""
    form = TransactionForm()
    form.tag_ids.choices = _tag_choices()

    if form.validate_on_submit():
        transaction = Transaction(
            type=form.type.data,
            amount=form.amount.data,
            date=form.date.data,
            description=form.description.data.strip(),
            category=(form.category.data or "").strip() or None,
            source=form.source.data,
            created_by_id=current_user.id,
        )
        if form.tag_ids.data:
            tags = db.session.scalars(db.select(Tag).where(Tag.id.in_(form.tag_ids.data))).all()
            transaction.tags = tags
        _apply_donor_fields(form, transaction)
        db.session.add(transaction)
        db.session.commit()
        flash("Transaction enregistrée.", "success")
        return redirect(url_for("treasury.detail", transaction_id=transaction.id))

    return render_template("treasury/form.html", form=form, title="Nouvelle transaction")


# ---------------------------------------------------------------------------
# Edit transaction — bureau only
# ---------------------------------------------------------------------------


@bp.route("/<int:transaction_id>/edit", methods=["GET", "POST"])
@bureau_required
def edit(transaction_id: int):
    """Edit an existing transaction."""
    transaction = db.session.get(
        Transaction, transaction_id, options=[selectinload(Transaction.tags)]
    )
    if transaction is None:
        abort(404)

    form = TransactionForm(obj=transaction)
    form.tag_ids.choices = _tag_choices()
    if request.method == "GET":
        form.tag_ids.data = [t.id for t in transaction.tags]

    if form.validate_on_submit():
        transaction.type = form.type.data
        transaction.amount = form.amount.data
        transaction.date = form.date.data
        transaction.description = form.description.data.strip()
        transaction.category = (form.category.data or "").strip() or None
        transaction.source = form.source.data
        tags = (
            db.session.scalars(db.select(Tag).where(Tag.id.in_(form.tag_ids.data))).all()
            if form.tag_ids.data
            else []
        )
        transaction.tags = list(tags)
        _apply_donor_fields(form, transaction)
        db.session.commit()
        flash("Transaction mise à jour.", "success")
        return redirect(url_for("treasury.detail", transaction_id=transaction.id))

    return render_template(
        "treasury/form.html", form=form, title="Modifier la transaction", transaction=transaction
    )


# ---------------------------------------------------------------------------
# Delete transaction — bureau only
# ---------------------------------------------------------------------------


@bp.route("/<int:transaction_id>/delete", methods=["POST"])
@bureau_required
def delete(transaction_id: int):
    """Delete a financial transaction and redirect to the list."""
    transaction = db.session.get(Transaction, transaction_id)
    if transaction is None:
        abort(404)

    db.session.delete(transaction)
    db.session.commit()
    flash("Transaction supprimée.", "success")
    return redirect(url_for("treasury.list_transactions"))


# ---------------------------------------------------------------------------
# CERFA generation — bureau only
# ---------------------------------------------------------------------------


@bp.route("/<int:transaction_id>/cerfa")
@bureau_required
def generate_cerfa(transaction_id: int):
    """Generate and serve the filled CERFA DOCX for a donation transaction."""
    transaction = db.session.get(Transaction, transaction_id)
    if transaction is None:
        abort(404)
    if transaction.source != TransactionSource.DONATION.value:
        flash("Le CERFA n'est disponible que pour les transactions de type Don.", "warning")
        return redirect(url_for("treasury.detail", transaction_id=transaction_id))

    cfg = AssociationConfig.get()
    if not cfg.name:
        flash(
            "Configurez d'abord les informations de l'association avant de générer un CERFA.",
            "warning",
        )
        return redirect(url_for("treasury.association_config"))

    from app.services.cerfa import generate_cerfa_pdf

    try:
        pdf_bytes, pdf_filename = generate_cerfa_pdf(transaction, cfg)
    except Exception as exc:
        flash(f"Erreur lors de la génération du CERFA : {exc}", "danger")
        return redirect(url_for("treasury.detail", transaction_id=transaction_id))

    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{pdf_filename}"'},
    )


@bp.route("/<int:transaction_id>/cerfa/send", methods=["POST"])
@bureau_required
def send_cerfa_email(transaction_id: int):
    """Queue CERFA generation + email delivery as an async Celery task."""
    transaction = db.session.get(Transaction, transaction_id)
    if transaction is None:
        abort(404)
    if transaction.source != TransactionSource.DONATION.value:
        abort(400)
    if not transaction.donor_email:
        flash("Aucun email de donateur enregistré pour cette transaction.", "warning")
        return redirect(url_for("treasury.detail", transaction_id=transaction_id))

    cfg = AssociationConfig.get()
    if not cfg.name:
        flash("Configurez d'abord les informations de l'association.", "warning")
        return redirect(url_for("treasury.association_config"))

    from app.tasks.cerfa import generate_and_send_cerfa

    generate_and_send_cerfa.delay(transaction_id)
    flash(
        f"Génération et envoi du CERFA à {transaction.donor_email} en cours…",
        "info",
    )
    return redirect(url_for("treasury.detail", transaction_id=transaction_id))


# ---------------------------------------------------------------------------
# Association configuration — bureau only
# ---------------------------------------------------------------------------


@bp.route("/config", methods=["GET", "POST"])
@bureau_required
def association_config():
    """Edit the association's legal identity used for CERFA generation."""
    cfg = AssociationConfig.get()
    form = AssociationConfigForm(obj=cfg)

    if form.validate_on_submit():
        form.populate_obj(cfg)
        db.session.commit()
        flash("Configuration mise à jour.", "success")
        return redirect(url_for("treasury.association_config"))

    return render_template("treasury/config.html", form=form, cfg=cfg)


# Only the in-kind ("don en nature") receipt uses a DOCX letter template;
# particulier/entreprise receipts are generated from the official CERFA AcroForm.
_CERFA_COLUMN_MAP = {
    "nature": "cerfa_tpl_nature",
}
_CERFA_LABELS = {
    "nature": "Don en nature",
}


@bp.route("/config/cerfa/<tpl_type>/upload", methods=["POST"])
@bureau_required
def upload_cerfa_template(tpl_type: str):
    """Upload a CERFA DOCX template."""
    col = _CERFA_COLUMN_MAP.get(tpl_type)
    if col is None:
        abort(404)

    file = request.files.get("file")
    if not file or not file.filename:
        flash("Aucun fichier sélectionné.", "warning")
        return redirect(url_for("treasury.association_config"))

    if not file.filename.lower().endswith(".docx"):
        flash("Seuls les fichiers DOCX sont acceptés.", "danger")
        return redirect(url_for("treasury.association_config"))

    data = file.read()
    if len(data) > 5 * 1024 * 1024:
        flash("Le fichier dépasse 5 Mo.", "danger")
        return redirect(url_for("treasury.association_config"))

    # Validate merge fields (in-kind letter: no "Montant_perçu" — uses "Don")
    required_fields = {"IDRECU", "Date", "Nom", "Don"}
    missing = _check_merge_fields(data, required_fields)
    if missing:
        flash(
            f"Champs de fusion manquants dans le modèle : {', '.join(sorted(missing))}. "
            "Le fichier n'a pas été enregistré.",
            "danger",
        )
        return redirect(url_for("treasury.association_config"))

    cfg = AssociationConfig.get()
    setattr(cfg, col, data)
    db.session.commit()
    label = _CERFA_LABELS.get(tpl_type, tpl_type)
    flash(f"Modèle « {label} » téléversé.", "success")
    return redirect(url_for("treasury.association_config"))


@bp.route("/config/cerfa/<tpl_type>/download")
@bureau_required
def download_cerfa_template(tpl_type: str):
    """Download the stored CERFA DOCX template."""
    col = _CERFA_COLUMN_MAP.get(tpl_type)
    if col is None:
        abort(404)

    cfg = AssociationConfig.get()
    data = getattr(cfg, col)
    if data is None:
        flash("Aucun modèle enregistré.", "warning")
        return redirect(url_for("treasury.association_config"))

    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="cerfa_{tpl_type}.docx"'},
    )


@bp.route("/config/cerfa/<tpl_type>/delete", methods=["POST"])
@bureau_required
def delete_cerfa_template(tpl_type: str):
    """Remove a stored CERFA template."""
    col = _CERFA_COLUMN_MAP.get(tpl_type)
    if col is None:
        abort(404)

    cfg = AssociationConfig.get()
    setattr(cfg, col, None)
    db.session.commit()
    label = _CERFA_LABELS.get(tpl_type, tpl_type)
    flash(f"Modèle « {label} » supprimé.", "info")
    return redirect(url_for("treasury.association_config"))


@bp.route("/config/logo/upload", methods=["POST"])
@bureau_required
def upload_logo():
    """Upload the association logo (PNG/JPG) for certificates."""
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Aucun fichier sélectionné.", "warning")
        return redirect(url_for("treasury.association_config"))

    if file.filename.lower().rsplit(".", 1)[-1] not in ("png", "jpg", "jpeg", "webp"):
        flash("Seuls les fichiers PNG, JPG ou WebP sont acceptés.", "danger")
        return redirect(url_for("treasury.association_config"))

    data = file.read()
    if len(data) > 2 * 1024 * 1024:
        flash("Le fichier dépasse 2 Mo.", "danger")
        return redirect(url_for("treasury.association_config"))

    cfg = AssociationConfig.get()
    cfg.logo = data
    db.session.commit()
    flash("Logo téléversé.", "success")
    return redirect(url_for("treasury.association_config"))


@bp.route("/config/logo/delete", methods=["POST"])
@bureau_required
def delete_logo():
    """Remove the stored association logo."""
    cfg = AssociationConfig.get()
    cfg.logo = None
    db.session.commit()
    flash("Logo supprimé.", "info")
    return redirect(url_for("treasury.association_config"))


@bp.route("/config/signature/upload", methods=["POST"])
@bureau_required
def upload_signature():
    """Upload the representative's signature image stamped on CERFA receipts."""
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Aucun fichier sélectionné.", "warning")
        return redirect(url_for("treasury.association_config"))

    if file.filename.lower().rsplit(".", 1)[-1] not in ("png", "jpg", "jpeg", "webp"):
        flash("Seuls les fichiers PNG, JPG ou WebP sont acceptés.", "danger")
        return redirect(url_for("treasury.association_config"))

    data = file.read()
    if len(data) > 2 * 1024 * 1024:
        flash("Le fichier dépasse 2 Mo.", "danger")
        return redirect(url_for("treasury.association_config"))

    cfg = AssociationConfig.get()
    cfg.signature = data
    db.session.commit()
    flash("Signature téléversée.", "success")
    return redirect(url_for("treasury.association_config"))


@bp.route("/config/signature/delete", methods=["POST"])
@bureau_required
def delete_signature():
    """Remove the stored signature image."""
    cfg = AssociationConfig.get()
    cfg.signature = None
    db.session.commit()
    flash("Signature supprimée.", "info")
    return redirect(url_for("treasury.association_config"))


# ---------------------------------------------------------------------------
# Excel export — bureau only
# ---------------------------------------------------------------------------


@bp.route("/export.xlsx")
@bureau_required
def export_xlsx():
    """Export all transactions as an Excel workbook (.xlsx)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    transactions = db.session.scalars(
        db.select(Transaction)
        .options(selectinload(Transaction.tags), selectinload(Transaction.created_by))
        .order_by(Transaction.date.desc(), Transaction.created_at.desc())
    ).all()

    wb = Workbook()

    # ---- Sheet 1: all transactions ----
    ws = wb.active
    ws.title = "Transactions"
    headers = [
        "Date",
        "Type",
        "Montant (€)",
        "Description",
        "Catégorie",
        "Source",
        "Saisie par",
        "Étiquettes",
    ]
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    source_labels = {
        "manual": "Saisie manuelle",
        "event": "Événement",
        "expense": "Note de frais",
        "donation": "Don",
        "membership": "Adhésion",
    }
    for row_num, tx in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1, value=tx.date.strftime("%d/%m/%Y"))
        ws.cell(row=row_num, column=2, value="Recette" if tx.type == "income" else "Dépense")
        amount_cell = ws.cell(row=row_num, column=3, value=float(tx.signed_amount))
        amount_cell.number_format = "#,##0.00 €"
        if tx.type == "income":
            amount_cell.font = Font(color="16A34A")
        else:
            amount_cell.font = Font(color="DC2626")
        ws.cell(row=row_num, column=4, value=tx.description)
        ws.cell(row=row_num, column=5, value=tx.category or "")
        ws.cell(row=row_num, column=6, value=source_labels.get(tx.source, tx.source))
        ws.cell(row=row_num, column=7, value=tx.created_by.full_name)
        ws.cell(row=row_num, column=8, value=", ".join(t.label for t in tx.tags))

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # ---- Sheet 2: summary ----
    ws2 = wb.create_sheet("Récapitulatif")
    total_income = sum(float(t.amount) for t in transactions if t.type == "income")
    total_expense = sum(float(t.amount) for t in transactions if t.type == "expense")
    ws2.append(["Recettes totales", total_income])
    ws2.append(["Dépenses totales", total_expense])
    ws2.append(["Solde", total_income - total_expense])
    for row in ws2.iter_rows(min_row=1, max_row=3):
        row[0].font = Font(bold=True)
        row[1].number_format = "#,##0.00 €"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="tresorerie.xlsx"'},
    )


# ---------------------------------------------------------------------------
# Tags management — bureau only
# ---------------------------------------------------------------------------


@bp.route("/cerfa-stats")
@bureau_required
def cerfa_stats():
    """CERFA summary: count and amounts by year and donation type."""
    from sqlalchemy import extract, func

    rows = db.session.execute(
        db.select(
            extract("year", Transaction.date).label("year"),
            Transaction.donation_type,
            func.count().label("count"),
            func.sum(Transaction.amount).label("total"),
        )
        .where(
            Transaction.source == TransactionSource.DONATION.value,
            Transaction.cerfa_sent_at.is_not(None),
        )
        .group_by("year", Transaction.donation_type)
        .order_by(extract("year", Transaction.date).desc(), Transaction.donation_type)
    ).all()

    # Group by year
    years: dict = {}
    for year, dtype, count, total in rows:
        y = int(year)
        if y not in years:
            years[y] = {"categories": {}, "total_count": 0, "total_amount": Decimal("0")}
        labels = {
            "particulier": "Don de particulier",
            "mecena": "Mécénat entreprise",
            "nature": "Don en nature",
        }
        label = labels.get(dtype or "particulier", dtype or "particulier")
        years[y]["categories"][label] = {"count": count, "amount": total or Decimal("0")}
        years[y]["total_count"] += count
        years[y]["total_amount"] += total or Decimal("0")

    return render_template("treasury/cerfa_stats.html", years=years)


@bp.route("/tags", methods=["GET", "POST"])
@bureau_required
def manage_tags():
    """List tags and create new ones."""
    form = TagForm()

    if form.validate_on_submit():
        existing = db.session.execute(
            db.select(Tag).where(Tag.label == form.label.data.strip())
        ).scalar_one_or_none()
        if existing:
            flash(f"L'étiquette « {form.label.data} » existe déjà.", "warning")
        else:
            tag = Tag(label=form.label.data.strip(), color=form.color.data)
            db.session.add(tag)
            db.session.commit()
            flash(f"Étiquette « {tag.label} » créée.", "success")
        return redirect(url_for("treasury.manage_tags"))

    tags = db.session.scalars(db.select(Tag).order_by(Tag.label)).all()
    return render_template("treasury/tags.html", tags=tags, form=form)


@bp.route("/tags/<int:tag_id>/delete", methods=["POST"])
@bureau_required
def delete_tag(tag_id: int):
    """Delete a tag (does not affect the transactions it was applied to)."""
    tag = db.session.get(Tag, tag_id)
    if tag is None:
        abort(404)
    label = tag.label
    db.session.delete(tag)
    db.session.commit()
    flash(f"Étiquette « {label} » supprimée.", "success")
    return redirect(url_for("treasury.manage_tags"))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _tag_choices() -> list[tuple[int, str]]:
    """Return (id, label) for all tags, sorted alphabetically."""
    tags = db.session.scalars(db.select(Tag).order_by(Tag.label)).all()
    return [(t.id, t.label) for t in tags]


def _apply_donor_fields(form: TransactionForm, transaction: Transaction) -> None:
    """Copy donor fields from form to transaction (only relevant for donations)."""
    if form.source.data == TransactionSource.DONATION.value:
        transaction.donation_type = form.donation_type.data or "particulier"
        transaction.donor_first_name = (form.donor_first_name.data or "").strip() or None
        transaction.donor_name = (form.donor_name.data or "").strip() or None
        transaction.donor_address = (form.donor_address.data or "").strip() or None
        transaction.donor_zip = (form.donor_zip.data or "").strip() or None
        transaction.donor_city = (form.donor_city.data or "").strip() or None
        transaction.donor_email = (form.donor_email.data or "").strip() or None
        transaction.donor_description = (form.donor_description.data or "").strip() or None
    else:
        transaction.donation_type = None
        transaction.donor_first_name = None
        transaction.donor_name = None
        transaction.donor_address = None
        transaction.donor_zip = None
        transaction.donor_city = None
        transaction.donor_email = None
        transaction.donor_description = None


def _check_merge_fields(docx_bytes: bytes, required: set[str]) -> set[str]:
    """Return the set of required merge fields missing from a DOCX template."""
    import tempfile

    try:
        from mailmerge import MailMerge

        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
            tmp.write(docx_bytes)
            tmp_path = tmp.name
        try:
            with MailMerge(tmp_path) as doc:
                found = set(doc.get_merge_fields())
        finally:
            import os

            os.unlink(tmp_path)
        return required - found
    except Exception:
        logger.warning("Could not inspect merge fields in uploaded DOCX")
        return set()  # allow upload if inspection fails
